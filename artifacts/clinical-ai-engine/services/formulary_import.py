"""
Import the hospital formulary, and refuse anything that fails the safety
contract.

This is the only way clinical numbers enter the system. The rules it enforces
are the same ones stated in `models/formulary.py` and backed by CHECK
constraints in migration 0003 — deliberately repeated here so a rejection can
name the row, the field and the reason, which a constraint violation cannot.

Two properties matter as much as the validation:

  * **Idempotent.** A file is identified by its SHA-256. Re-importing the same
    file changes nothing, so an operator who is unsure whether an import ran can
    simply run it again.
  * **A changed value resets approval.** If an import alters any clinical figure
    on an already-approved drug, that drug returns to `pending` and stops
    producing a dose until a pharmacist signs off the new number. Inheriting an
    old approval across a changed dose is the failure this exists to prevent.
"""
import csv
import hashlib
import io
import logging
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


# ── The canonical column set ──────────────────────────────────────────────────
# A hospital's export will not use these headers. `config/formulary_mapping.yaml`
# maps their column names onto these; see load_mapping().

TEXT_FIELDS = (
    "generic_name", "name_ar", "unit", "frequency", "route", "antidote",
    "reference_regimen", "source_name", "source_edition", "source_ref",
)
NUMERIC_FIELDS = (
    "dose_per_kg", "adult_flat_min", "adult_flat_max", "adult_max_dose",
    "adult_max_daily", "adult_max_daily_elderly", "pediatric_min_per_kg",
    "pediatric_max_per_kg", "pediatric_max_daily_per_kg",
    "overdose_threshold_absolute", "overdose_threshold_per_kg",
)
INT_FIELDS = ("pediatric_max_doses_per_day",)
LIST_FIELDS = ("aliases", "contraindications", "interactions", "warnings")
BOOL_FIELDS = ("high_risk", "auto_calculate")

ALL_FIELDS = TEXT_FIELDS + NUMERIC_FIELDS + INT_FIELDS + LIST_FIELDS + BOOL_FIELDS

# Without these a row is not importable at all. `source_name` and `source_ref`
# are here because a dose with no citable origin is precisely what this whole
# mechanism exists to stop.
REQUIRED_FIELDS = ("generic_name", "unit", "source_name", "source_ref")

VALID_UNITS = ("mg", "unit")

# The fields that carry a milligram figure. A drug dosed in international units
# must set none of them.
MG_FIELDS = NUMERIC_FIELDS

# Changing any of these is a clinical change and invalidates a prior approval.
CLINICAL_FIELDS = NUMERIC_FIELDS + INT_FIELDS + (
    "unit", "auto_calculate", "frequency", "route", "antidote",
    "reference_regimen", "contraindications", "interactions", "warnings",
    "high_risk",
)


class FormularyFileError(ValueError):
    """The file itself could not be read — not a per-row problem."""


@dataclass
class RowOutcome:
    row_number: int
    generic_name: str
    action: str  # inserted | updated | unchanged | rejected
    reason: Optional[str] = None
    changed_fields: List[str] = field(default_factory=list)

    @property
    def rejected(self) -> bool:
        return self.action == "rejected"


@dataclass
class ImportReport:
    file_name: str
    sha256: str
    dry_run: bool
    already_imported: bool = False
    unknown_columns: List[str] = field(default_factory=list)
    missing_columns: List[str] = field(default_factory=list)
    rows: List[RowOutcome] = field(default_factory=list)

    def tally(self, action: str) -> int:
        return sum(1 for r in self.rows if r.action == action)

    @property
    def rejected(self) -> List[RowOutcome]:
        return [r for r in self.rows if r.rejected]

    def to_dict(self) -> Dict[str, Any]:
        return {
            "file_name": self.file_name,
            "sha256": self.sha256,
            "dry_run": self.dry_run,
            "already_imported": self.already_imported,
            "unknown_columns": self.unknown_columns,
            "missing_columns": self.missing_columns,
            "summary": {
                "total": len(self.rows),
                "inserted": self.tally("inserted"),
                "updated": self.tally("updated"),
                "unchanged": self.tally("unchanged"),
                "rejected": self.tally("rejected"),
            },
            "rows": [
                {
                    "row": r.row_number,
                    "drug": r.generic_name,
                    "action": r.action,
                    "reason": r.reason,
                    "changed_fields": r.changed_fields,
                }
                for r in self.rows
            ],
        }


# ── Reading the file ──────────────────────────────────────────────────────────

def load_mapping(text: str) -> Dict[str, str]:
    """
    Parse a column mapping: `canonical_field: Source Column Header`, one per
    line. Deliberately not YAML — this is a flat string-to-string map and a
    parser dependency buys nothing.
    """
    mapping: Dict[str, str] = {}
    for raw in text.splitlines():
        line = raw.split("#", 1)[0].strip()
        if not line or ":" not in line:
            continue
        canonical, source = line.split(":", 1)
        canonical = canonical.strip()
        source = source.strip().strip('"').strip("'")
        if canonical in ALL_FIELDS and source:
            mapping[canonical] = source
    return mapping


def read_table(content: bytes, file_name: str) -> Tuple[List[str], List[List[Any]]]:
    """Return (headers, rows) from a CSV or XLSX file."""
    lowered = file_name.lower()

    if lowered.endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except ImportError as e:  # pragma: no cover - dependency is declared
            raise FormularyFileError(f"openpyxl is required to read {file_name}: {e}")
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]]
        rows = [list(r) for r in sheet.iter_rows(values_only=True)]
        wb.close()
        if not rows:
            raise FormularyFileError("The spreadsheet is empty")
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        return headers, rows[1:]

    if lowered.endswith((".csv", ".tsv", ".txt")):
        delimiter = "\t" if lowered.endswith(".tsv") else ","
        text = content.decode("utf-8-sig", errors="replace")
        reader = list(csv.reader(io.StringIO(text), delimiter=delimiter))
        if not reader:
            raise FormularyFileError("The file is empty")
        return [h.strip() for h in reader[0]], reader[1:]

    raise FormularyFileError(
        f"Unsupported file type: {file_name}. Provide .csv, .tsv or .xlsx."
    )


# ── Coercion ──────────────────────────────────────────────────────────────────

def _text(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _number(value: Any, field_name: str) -> Tuple[Optional[float], Optional[str]]:
    s = _text(value)
    if s is None:
        return None, None
    s = s.replace(",", "")
    try:
        number = float(s)
    except ValueError:
        return None, f"{field_name}: {s!r} is not a number"
    if number < 0:
        return None, f"{field_name}: {number} is negative"
    return number, None


def _list(value: Any) -> List[str]:
    s = _text(value)
    if not s:
        return []
    parts = s.split("|") if "|" in s else s.split(";")
    return [p.strip() for p in parts if p.strip()]


def _bool(value: Any, default: bool) -> bool:
    s = _text(value)
    if s is None:
        return default
    return s.lower() in ("1", "true", "yes", "y", "t", "نعم")


# ── Validation ────────────────────────────────────────────────────────────────

def validate(raw: Dict[str, Any]) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    """
    Turn one mapped row into a database-ready dict, or explain the refusal.

    Every rule here corresponds to a defect that reached production once.
    """
    row: Dict[str, Any] = {}

    for f in TEXT_FIELDS:
        row[f] = _text(raw.get(f))
    for f in LIST_FIELDS:
        row[f] = _list(raw.get(f))
    for f in NUMERIC_FIELDS:
        value, err = _number(raw.get(f), f)
        if err:
            return None, err
        row[f] = value
    for f in INT_FIELDS:
        value, err = _number(raw.get(f), f)
        if err:
            return None, err
        row[f] = int(value) if value is not None else None

    for f in REQUIRED_FIELDS:
        if not row.get(f):
            return None, f"{f} is required"

    row["generic_name"] = row["generic_name"].lower()
    row["unit"] = row["unit"].lower()
    if row["unit"] not in VALID_UNITS:
        return None, (
            f"unit: {row['unit']!r} is not one of {VALID_UNITS}. Every numeric "
            "field must be expressed in a declared unit."
        )

    row["high_risk"] = _bool(raw.get("high_risk"), False)
    row["auto_calculate"] = _bool(raw.get("auto_calculate"), True)

    # A drug dosed in international units is never computed, and must not carry
    # milligram figures that would imply otherwise. Heparin was modelled in mg.
    if row["unit"] != "mg":
        row["auto_calculate"] = False
        populated = [f for f in MG_FIELDS if row.get(f) is not None]
        if populated:
            return None, (
                f"unit is {row['unit']!r} but milligram fields are set: "
                f"{', '.join(populated)}. Put the regimen in reference_regimen."
            )
        if not row.get("reference_regimen"):
            return None, (
                "unit-dosed drugs need reference_regimen — with no computed "
                "number and no regimen text there is nothing to show a nurse"
            )

    # Paired bounds. A lone bound silently disables the range it belongs to.
    for lo, hi in (
        ("adult_flat_min", "adult_flat_max"),
        ("pediatric_min_per_kg", "pediatric_max_per_kg"),
    ):
        if (row.get(lo) is None) != (row.get(hi) is None):
            return None, f"{lo} and {hi} must both be set, or neither"
        if row.get(lo) is not None and row[lo] > row[hi]:
            return None, f"{lo} ({row[lo]}) is greater than {hi} ({row[hi]})"

    # The enoxaparin defect: a per-kg value parked in the absolute field read a
    # normal 70 mg dose as a 35x overdose requiring protamine.
    absolute = row.get("overdose_threshold_absolute")
    if absolute is not None:
        if row.get("adult_max_daily") is not None and absolute < row["adult_max_daily"]:
            return None, (
                f"overdose_threshold_absolute ({absolute}) is below "
                f"adult_max_daily ({row['adult_max_daily']}) — this is an "
                "absolute total, not a per-kg value. Use "
                "overdose_threshold_per_kg for weight-scaled limits."
            )
        if row.get("adult_max_dose") is not None and absolute < row["adult_max_dose"]:
            return None, (
                f"overdose_threshold_absolute ({absolute}) is below "
                f"adult_max_dose ({row['adult_max_dose']})"
            )

    # A per-kg ceiling below the per-kg dose flags every therapeutic dose.
    per_kg = row.get("overdose_threshold_per_kg")
    if per_kg is not None and row.get("dose_per_kg") is not None:
        if per_kg < row["dose_per_kg"]:
            return None, (
                f"overdose_threshold_per_kg ({per_kg}) is below dose_per_kg "
                f"({row['dose_per_kg']}) — every therapeutic dose would be "
                "flagged as an overdose"
            )

    if row.get("adult_max_dose") is not None and row.get("adult_max_daily") is not None:
        if row["adult_max_dose"] > row["adult_max_daily"]:
            return None, (
                f"adult_max_dose ({row['adult_max_dose']}) exceeds "
                f"adult_max_daily ({row['adult_max_daily']})"
            )

    # An auto-calculating entry that can produce no number at all is a row that
    # will always answer "per physician order" — accept it, but it must at least
    # say something.
    if row["auto_calculate"] and not any(
        row.get(f) is not None
        for f in ("dose_per_kg", "adult_flat_min", "adult_max_daily")
    ):
        if not row.get("reference_regimen"):
            return None, (
                "no dose figures and no reference_regimen — this row cannot "
                "tell a nurse anything"
            )

    return row, None


# ── Importing ─────────────────────────────────────────────────────────────────

def file_digest(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _differences(existing: Dict[str, Any], incoming: Dict[str, Any]) -> List[str]:
    """Clinical fields whose value changed. Comparison is on normalised types."""
    changed = []
    for f in CLINICAL_FIELDS:
        old, new = existing.get(f), incoming.get(f)
        if isinstance(old, list) or isinstance(new, list):
            if sorted(old or []) != sorted(new or []):
                changed.append(f)
            continue
        if old is None and new is None:
            continue
        if old is None or new is None:
            changed.append(f)
            continue
        if isinstance(old, (int, float)) and isinstance(new, (int, float)):
            if abs(float(old) - float(new)) > 1e-9:
                changed.append(f)
            continue
        if str(old) != str(new):
            changed.append(f)
    return changed


SELECT_EXISTING = """
    SELECT drug_id, generic_name, unit, auto_calculate, dose_per_kg,
           adult_flat_min, adult_flat_max, adult_max_dose, adult_max_daily,
           adult_max_daily_elderly, pediatric_min_per_kg, pediatric_max_per_kg,
           pediatric_max_daily_per_kg, pediatric_max_doses_per_day,
           overdose_threshold_absolute, overdose_threshold_per_kg,
           frequency, route, antidote, reference_regimen,
           contraindications, interactions, warnings, high_risk,
           review_status, version
    FROM bnp_drug_formulary
    WHERE retired_at IS NULL
"""

INSERT_SQL = """
    INSERT INTO bnp_drug_formulary (
        generic_name, name_ar, aliases, unit, auto_calculate, dose_per_kg,
        adult_flat_min, adult_flat_max, adult_max_dose, adult_max_daily,
        adult_max_daily_elderly, pediatric_min_per_kg, pediatric_max_per_kg,
        pediatric_max_daily_per_kg, pediatric_max_doses_per_day,
        overdose_threshold_absolute, overdose_threshold_per_kg,
        frequency, route, antidote, reference_regimen,
        contraindications, interactions, warnings, high_risk,
        source_name, source_edition, source_ref,
        imported_from_file, imported_file_sha256, imported_at, imported_by,
        review_status, version
    ) VALUES (
        %(generic_name)s, %(name_ar)s, %(aliases)s, %(unit)s, %(auto_calculate)s,
        %(dose_per_kg)s, %(adult_flat_min)s, %(adult_flat_max)s,
        %(adult_max_dose)s, %(adult_max_daily)s, %(adult_max_daily_elderly)s,
        %(pediatric_min_per_kg)s, %(pediatric_max_per_kg)s,
        %(pediatric_max_daily_per_kg)s, %(pediatric_max_doses_per_day)s,
        %(overdose_threshold_absolute)s, %(overdose_threshold_per_kg)s,
        %(frequency)s, %(route)s, %(antidote)s, %(reference_regimen)s,
        %(contraindications)s, %(interactions)s, %(warnings)s, %(high_risk)s,
        %(source_name)s, %(source_edition)s, %(source_ref)s,
        %(file_name)s, %(sha256)s, %(now)s, %(actor)s,
        'pending', 1
    )
"""

UPDATE_SQL = """
    UPDATE bnp_drug_formulary SET
        name_ar = %(name_ar)s, aliases = %(aliases)s, unit = %(unit)s,
        auto_calculate = %(auto_calculate)s, dose_per_kg = %(dose_per_kg)s,
        adult_flat_min = %(adult_flat_min)s, adult_flat_max = %(adult_flat_max)s,
        adult_max_dose = %(adult_max_dose)s, adult_max_daily = %(adult_max_daily)s,
        adult_max_daily_elderly = %(adult_max_daily_elderly)s,
        pediatric_min_per_kg = %(pediatric_min_per_kg)s,
        pediatric_max_per_kg = %(pediatric_max_per_kg)s,
        pediatric_max_daily_per_kg = %(pediatric_max_daily_per_kg)s,
        pediatric_max_doses_per_day = %(pediatric_max_doses_per_day)s,
        overdose_threshold_absolute = %(overdose_threshold_absolute)s,
        overdose_threshold_per_kg = %(overdose_threshold_per_kg)s,
        frequency = %(frequency)s, route = %(route)s, antidote = %(antidote)s,
        reference_regimen = %(reference_regimen)s,
        contraindications = %(contraindications)s,
        interactions = %(interactions)s, warnings = %(warnings)s,
        high_risk = %(high_risk)s,
        source_name = %(source_name)s, source_edition = %(source_edition)s,
        source_ref = %(source_ref)s,
        imported_from_file = %(file_name)s, imported_file_sha256 = %(sha256)s,
        imported_at = %(now)s, imported_by = %(actor)s,
        -- A changed clinical value invalidates the previous sign-off. Carrying
        -- an approval across an altered dose is the failure this prevents.
        review_status = 'pending', reviewed_by = NULL, reviewer_license = NULL,
        reviewed_at = NULL, review_note = NULL,
        version = version + 1, updated_at = %(now)s
    WHERE drug_id = %(drug_id)s
"""


def _params(row: Dict[str, Any], file_name: str, sha256: str, actor: str) -> Dict:
    from psycopg2.extras import Json

    params = dict(row)
    for f in LIST_FIELDS:
        params[f] = Json(row.get(f) or [])
    params.update(
        file_name=file_name, sha256=sha256, actor=actor, now=datetime.utcnow()
    )
    return params


def import_formulary(
    *,
    content: bytes,
    file_name: str,
    mapping: Dict[str, str],
    actor: str,
    dry_run: bool = True,
) -> ImportReport:
    """
    Validate a formulary file and, unless this is a dry run, apply it.

    Dry run is the default: an import that silently rewrote clinical figures on
    a first attempt would be the wrong shape of tool.
    """
    from models.database import db_cursor

    sha256 = file_digest(content)
    headers, raw_rows = read_table(content, file_name)

    report = ImportReport(file_name=file_name, sha256=sha256, dry_run=dry_run)

    # Which canonical field does each column carry? Unmapped columns are
    # reported rather than guessed at.
    header_index: Dict[str, int] = {}
    for canonical, source_header in mapping.items():
        if source_header in headers:
            header_index[canonical] = headers.index(source_header)
    for canonical in ALL_FIELDS:
        if canonical in header_index:
            continue
        if canonical in headers:  # file already uses our names
            header_index[canonical] = headers.index(canonical)

    mapped_headers = {headers[i] for i in header_index.values()}
    report.unknown_columns = [h for h in headers if h and h not in mapped_headers]
    report.missing_columns = [f for f in REQUIRED_FIELDS if f not in header_index]
    if report.missing_columns:
        return report

    with db_cursor() as (cur, _conn):
        cur.execute(
            "SELECT 1 FROM bnp_drug_formulary "
            "WHERE imported_file_sha256 = %s LIMIT 1",
            (sha256,),
        )
        if cur.fetchone():
            report.already_imported = True
            return report

        cur.execute(SELECT_EXISTING)
        existing = {r["generic_name"]: dict(r) for r in cur.fetchall()}

        for offset, raw_row in enumerate(raw_rows):
            row_number = offset + 2  # 1-based, plus the header line
            mapped = {
                canonical: (raw_row[i] if i < len(raw_row) else None)
                for canonical, i in header_index.items()
            }
            name_for_report = str(mapped.get("generic_name") or "").strip() or "(blank)"

            if not any(str(v).strip() for v in mapped.values() if v is not None):
                continue  # blank line

            row, error = validate(mapped)
            if error:
                report.rows.append(
                    RowOutcome(row_number, name_for_report, "rejected", error)
                )
                continue

            prior = existing.get(row["generic_name"])
            params = _params(row, file_name, sha256, actor)

            if prior is None:
                report.rows.append(
                    RowOutcome(row_number, row["generic_name"], "inserted")
                )
                if not dry_run:
                    cur.execute(INSERT_SQL, params)
                continue

            changed = _differences(
                {k: _decimal_to_float(v) for k, v in prior.items()}, row
            )
            if not changed:
                report.rows.append(
                    RowOutcome(row_number, row["generic_name"], "unchanged")
                )
                continue

            report.rows.append(
                RowOutcome(
                    row_number,
                    row["generic_name"],
                    "updated",
                    reason=(
                        "clinical values changed — approval reset to pending"
                        if prior.get("review_status") == "approved"
                        else None
                    ),
                    changed_fields=changed,
                )
            )
            if not dry_run:
                params["drug_id"] = prior["drug_id"]
                cur.execute(UPDATE_SQL, params)

    logger.info(
        "Formulary import %s: %s — %d inserted, %d updated, %d unchanged, "
        "%d rejected",
        "validated" if dry_run else "applied",
        file_name,
        report.tally("inserted"),
        report.tally("updated"),
        report.tally("unchanged"),
        report.tally("rejected"),
    )
    return report


def _decimal_to_float(value):
    from decimal import Decimal

    return float(value) if isinstance(value, Decimal) else value
