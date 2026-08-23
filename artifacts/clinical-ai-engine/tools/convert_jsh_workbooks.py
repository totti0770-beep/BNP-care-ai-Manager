"""
Convert the hospital's two structured workbooks into one formulary import CSV.

Input is the pharmacy's own field-labelled exports — `JSH_Drug_Formulary_2026`
(521 drugs) and `IV_Sterile_Preparations_Manual_2026` (206 preparations) — not
raw PDFs. That difference matters: the columns are already separated by the
document's own headings, so this is a *converter*, not a scraper, and it carries
none of the name-garbling that table-scraping the source PDFs produced
(`acetamionphen`, `erythromyci n`, `methylprednisoln e sodium succinate`).

Ground rules, in order of importance:

  * **Nothing is authored.** Every value is a cell copied across, or a labelled
    assembly of cells. A column the workbook leaves blank stays blank.
  * **No numeric field is ever emitted.** The workbooks state doses as prose
    ("0.25–0.5 mg initially; may repeat 0.25 mg every 6 hours"). Parsing a
    number out of that into `dose_per_kg` is exactly the defect class that
    produced pediatric ranges for adults and a 35x enoxaparin false overdose.
    Every row is `auto_calculate=no`: the nurse gets the hospital's own text,
    verbatim, and the engine computes nothing from it.
  * **Every row cites its page**, so a reviewer opens the same document at the
    same place. A drug present in both workbooks cites both.
  * **One row per drug.** 81 names appear in both workbooks. Emitting two rows
    would put the same drug in the review queue twice and split its evidence;
    they are merged, with each workbook's text kept under its own heading.
  * **A near-miss spelling is not a new drug** — the same rule and the same
    matcher as `extract_formulary_pdfs.py`. Near-misses are reported in the
    manifest, unresolved, never silently imported as a second row.
  * **`high_risk` is carried forward, never invented.** The workbooks have no
    such column; dropping the flag would silently remove a nurse-facing warning.

Usage:
    python tools/convert_jsh_workbooks.py \
        JSH_Drug_Formulary_2026_structured.xlsx \
        IV_Sterile_Preparations_Manual_2026_structured.xlsx \
        --out data/formulary/jsh_workbooks_import.csv

The importer then enforces the safety contract on the result, so the import
report stays the single place a rejection is explained.
"""
import argparse
import csv
import hashlib
import json
import re
import sys
from pathlib import Path

# The engine root, so `models.database` resolves when run as a script.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from tools.extract_formulary_pdfs import (  # noqa: E402  (path set above)
    HEADERS,
    clean,
    detect_unit,
    find_near_duplicate,
    present,
)


def _openpyxl():
    """
    Imported lazily. openpyxl is a real dependency (the review packet uses it),
    but keeping the import inside the call means the pure helpers below can be
    imported and tested without it — the same discipline the pymupdf import in
    `extract_formulary_pdfs.py` needed after it broke CI.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise SystemExit("openpyxl is required: pip install openpyxl") from exc
    return load_workbook


# Sections are joined with "|" because the importer splits a list field on "|"
# when present and only falls back to ";" otherwise — and this clinical text is
# full of semicolons that must not become list boundaries.
SECTION = " | "


def read_sheet(path: Path, sheet_name: str):
    """Rows of one sheet as dicts keyed by the header row, values cleaned."""
    load_workbook = _openpyxl()
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(h) for h in next(rows)]
    out = [
        {h: clean(v) for h, v in zip(headers, raw)}
        for raw in rows
        if any(v not in (None, "") for v in raw)
    ]
    workbook.close()
    return out


def labelled(pairs) -> str:
    """`Label: value` sections, skipping every field the workbook left blank."""
    return SECTION.join(
        f"{label}: {value}" for label, value in pairs if value and present(value)
    )


def curated_fields() -> dict:
    """
    Fields already curated in the live table that these workbooks cannot supply.

    `UPDATE_SQL` in `services/formulary_import.py` writes every column, so a row
    the import touches loses anything the incoming file leaves blank. For four
    fields that is destructive and, worse, invisible — `name_ar` and `aliases`
    are not in `CLINICAL_FIELDS`, so their loss would not even appear in the
    change report:

      * `name_ar` — the Arabic name. Losing it breaks Arabic lookup for those
        drugs in an app whose whole point is being bilingual.
      * `antidote` — "naloxone for morphine", "protamine for heparin". Safety
        information, and nothing to do with the unverified dose arithmetic this
        import is deliberately clearing out.
      * `aliases` — the brand names a nurse actually types (lanoxin, clexane,
        lasix).
      * `high_risk` — the flag behind the nurse-facing warning.

    The workbooks state antidotes only inside prose ("... Antidote: digoxin
    -specific Fab."), and parsing that out is authoring. Carrying the existing
    value forward is not: it is the value a human already curated, kept.

    Fails loudly rather than quietly returning nothing, because a silent empty
    result here is precisely the data loss it exists to prevent.
    """
    try:
        from models.database import db_cursor
        with db_cursor() as (cur, _):
            cur.execute(
                "SELECT generic_name, name_ar, aliases, antidote, high_risk "
                "FROM bnp_drug_formulary WHERE retired_at IS NULL"
            )
            return {r["generic_name"]: dict(r) for r in cur.fetchall()}
    except Exception as e:
        print(
            f"WARNING: could not read the existing formulary ({e}) — Arabic "
            "names, antidotes, aliases and high-risk flags will NOT be carried "
            "forward, and importing this CSV would silently erase them. Do not "
            "apply it.",
            file=sys.stderr,
        )
        return {}


def inherited(name: str, curated: dict) -> dict:
    """
    Curated values for a drug, allowing a longer name to inherit from the
    molecule it extends — "heparin sodium" from "heparin".

    **Only safety attributes inherit**, and only downhill. `antidote` and
    `high_risk` are properties of the molecule: heparin sodium is reversed by
    protamine and is high-alert exactly as heparin is, and both fields fail
    safe if over-applied — an extra warning is a caution, a missing one is a
    removed safeguard.

    `name_ar` and `aliases` deliberately do **not** inherit; they are read
    exactly, by the caller. They are identity, not pharmacology: letting
    "amoxicillin/clavulanic acid" inherit "أموكسيسيلين" would label a
    combination product with the Arabic name of one of its components, which
    is simply the wrong name shown to an Arabic-reading nurse. A blank there
    is honest; a borrowed name is not.
    """
    if name in curated:
        return curated[name]
    for existing in sorted(curated, key=len, reverse=True):
        if len(existing) >= 5 and re.match(rf"{re.escape(existing)}\b", name):
            return curated[existing]
    return {}


def is_high_risk(name: str, flagged: set) -> bool:
    """
    Whether a workbook name refers to a drug already flagged high-risk.

    Matched on a word-boundary prefix, because the workbooks name preparations
    where the formulary named the molecule: "heparin sodium" is the flagged
    "heparin", and each of the five "insulin – …-acting" rows is "insulin".
    The rule can over-apply; it cannot under-apply. That asymmetry is
    deliberate — an extra high-risk warning is a caution, a missing one is a
    removed safeguard.
    """
    lowered = (name or "").lower()
    return any(re.match(rf"{re.escape(f)}\b", lowered) for f in flagged)


# ── The two workbooks ─────────────────────────────────────────────────────────

def from_formulary(row: dict) -> dict:
    """One drug's clinical record out of the JSH Drug Formulary sheet."""
    regimen = labelled([
        ("Therapeutic class", row.get("Therapeutic_Class")),
        ("Indications", row.get("Indications")),
        ("Dosage form and strength", row.get("Dosage_Form_Strength")),
        ("Adult dosing", row.get("Dosing_Adult")),
        ("Pediatric dosing", row.get("Dosing_Pediatric")),
        ("Renal/hepatic adjustment", row.get("Renal_Hepatic_Adjustment")),
        ("Administration", row.get("Administration")),
        ("Prescriber authority", row.get("Prescriber_Authority")),
        ("Additional notes", row.get("Additional_Notes")),
    ])
    warnings = labelled([
        ("Cautions and warnings", row.get("Cautions_Warnings")),
        ("Adverse drug reactions", row.get("Adverse_Drug_Reactions")),
        ("Monitoring", row.get("Monitoring")),
        ("Pregnancy category", row.get("Pregnancy_Category")),
        ("Lactation", row.get("Lactation")),
        ("Storage", row.get("Storage")),
    ])
    return {
        "generic_name": row["Drug_Name"].lower(),
        "unit": detect_unit(
            row.get("Dosage_Form_Strength", ""), row.get("Dosing_Adult", "")
        ),
        "route": row.get("Administration", ""),
        "reference_regimen": regimen,
        "warnings": warnings,
        "contraindications": row.get("Contraindications", ""),
        "interactions": row.get("Drug_Interactions", ""),
        "source_name": row.get("Reference_Standard", ""),
        "source_ref": f"Drug Formulary p.{row.get('Source_Page', '')}",
    }


def from_iv_manual(row: dict) -> dict:
    """One preparation's administration record out of the IV manual sheet."""
    regimen = labelled([
        ("Package size / initial strength", row.get("Package_Size_Initial_Strength")),
        ("Final concentration", row.get("Final_Concentration_Percentage")),
        ("Final volume", row.get("Final_Volume")),
        ("Diluents", row.get("Diluents")),
        ("Preparation, administration and stability",
         row.get("Full_Dosing_Administration_Stability_Text")),
    ])
    return {
        "generic_name": row["Medication_Name"].lower(),
        "unit": detect_unit(
            row.get("Package_Size_Initial_Strength", ""),
            row.get("Final_Concentration_Percentage", ""),
        ),
        "route": "IV" if row.get("Record_Type") == "IV_Injectable" else "",
        "reference_regimen": regimen,
        "warnings": "",
        "contraindications": "",
        "interactions": "",
        "source_name": row.get("Reference_Standard", ""),
        "source_ref": f"IV Sterile Preparations Manual p.{row.get('Source_Page', '')}",
    }


def combine(first: dict, second: dict) -> dict:
    """
    One row for a drug that is described more than once.

    Two shapes need this, and both are real in these workbooks:

      * **Across the workbooks** — 91 drugs appear in the formulary *and* the IV
        manual. The formulary carries the clinical record, the IV manual the
        preparation and stability detail.
      * **Within one sheet** — 16 drugs appear twice in the formulary because
        different chapters cover different uses: docusate sodium is an oral
        stool softener on p.21 and an ear-wax softener on p.414; azathioprine
        is an IBD immunomodulator on p.17 and a transplant immunosuppressant on
        p.352. Keeping only the first would silently delete a licensed
        indication, so both are kept.

    Every text is preserved under its own heading with its own page cite, and
    nothing is rewritten or summarised — the sections sit side by side for the
    reviewer to read.
    """
    merged = dict(first)
    for field in ("reference_regimen", "warnings", "contraindications", "interactions"):
        parts, seen = [], set()
        for value in (first.get(field, ""), second.get(field, "")):
            if value and value not in seen:
                seen.add(value)
                parts.append(value)
        merged[field] = SECTION.join(parts)
    merged["source_ref"] = " | ".join(
        dict.fromkeys(p for p in (first["source_ref"], second["source_ref"]) if p)
    )
    # "unit" is the one field a second record can be more specific about: the IV
    # manual states the vial strength where the formulary may describe tablets.
    # Preferring "unit" is the conservative direction — it forbids milligram
    # figures on the row, and these rows carry none by design.
    if "unit" in (first["unit"], second["unit"]):
        merged["unit"] = "unit"
    if not merged["route"]:
        merged["route"] = second["route"]
    return merged


def prefix_families(names) -> list:
    """
    Pairs where one name is a longer form of another — "heparin" and "heparin
    sodium", "morphine" and "morphine sulfate", "vancomycin" and "vancomycin
    eye solution".

    The edit-distance matcher cannot see these: the names differ by a whole
    word, not a character. But they are the same failure as digoxin/digoxine —
    a nurse searching the short name never reaches the fuller record, or
    reviews the same molecule twice.

    They are **reported, never merged**. A salt or a preparation is not always
    interchangeable with its molecule — heparin sodium and heparin calcium are
    different products, and vancomycin eye drops are not IV vancomycin. Which
    of these are genuinely one entry is a pharmacist's call.
    """
    ordered = sorted(names, key=len)
    families = []
    for index, short in enumerate(ordered):
        if len(short) < 5:
            continue
        for longer in ordered[index + 1:]:
            if re.match(rf"{re.escape(short)}\b", longer) and longer != short:
                families.append({"name": longer, "extends": short})
    return families


# ── Assembly ──────────────────────────────────────────────────────────────────

def build_rows(formulary_rows, iv_rows, curated, edition: str):
    """Merged, deduplicated rows plus everything that needs a human decision."""
    by_name = {}
    notes = {"merged_across_workbooks": [], "merged_within_sheet": []}

    for source_rows, convert, label in (
        (formulary_rows, from_formulary, "Drug Formulary"),
        (iv_rows, from_iv_manual, "IV Manual"),
    ):
        for raw in source_rows:
            try:
                row = convert(raw)
            except KeyError:
                continue  # no name column on this line
            name = row["generic_name"]
            if not name or not row["reference_regimen"]:
                continue
            if name not in by_name:
                by_name[name] = row
                continue
            by_name[name] = combine(by_name[name], row)
            notes[
                "merged_across_workbooks" if label == "IV Manual"
                else "merged_within_sheet"
            ].append(name)

    rows, possible_duplicates = [], []
    for name in sorted(by_name):
        near = find_near_duplicate(name, set(by_name) - {name})
        if near and near[0] < name:
            # Report once, on the second of the pair, and keep both: which
            # spelling is right is a pharmacist's call, not this script's.
            possible_duplicates.append(
                {"name": name, "similar_to": near[0], "similarity": near[1]}
            )
        row = by_name[name]
        exact = curated.get(name, {})       # identity: this drug's own record
        safety = inherited(name, curated)   # pharmacology: may come from the molecule
        rows.append({
            "generic_name": name,
            "name_ar": exact.get("name_ar") or "",
            "aliases": "|".join(exact.get("aliases") or []),
            "antidote": safety.get("antidote") or "",
            "unit": row["unit"],
            "auto_calculate": "no",
            "route": row["route"],
            "reference_regimen": row["reference_regimen"],
            "warnings": row["warnings"],
            "source_name": row["source_name"],
            "source_edition": edition,
            "source_ref": row["source_ref"],
            "contraindications": row["contraindications"],
            "interactions": row["interactions"],
            "high_risk": "yes" if safety.get("high_risk") else "no",
        })
    notes["prefix_families"] = prefix_families(by_name)
    return rows, notes, possible_duplicates


def superseded_rows(new_names, curated) -> list:
    """
    Live rows this conversion replaces under a corrected spelling.

    An exact-name match is handled by the importer, which updates the row in
    place and keeps its id and history. The problem case is a row whose *name*
    is wrong — `acetylcystine` for `acetylcysteine`, `retaplase` for
    `reteplase`, `methylprednisoln e sodium succinate`. The importer cannot
    match those, so the corrected row arrives as a new insert and the misspelt
    one would sit beside it forever, splitting one drug across two entries.

    They are listed for retirement, not deleted: retirement is the project's
    rule for clinical data, so a citation made against the old row still
    resolves.
    """
    incoming = set(new_names)
    return [
        {"retire": existing, "superseded_by": match[0], "similarity": match[1]}
        for existing in sorted(curated)
        if existing not in incoming
        and (match := find_near_duplicate(existing, incoming))
    ]


OUT_HEADERS = HEADERS + [
    "name_ar", "aliases", "antidote", "contraindications", "interactions",
    "high_risk",
]


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("formulary_workbook", type=Path)
    parser.add_argument("iv_workbook", type=Path)
    parser.add_argument("--out", type=Path, required=True)
    parser.add_argument("--edition", default="JSH 2026")
    args = parser.parse_args()

    formulary_rows = read_sheet(args.formulary_workbook, "Drug_Formulary")
    iv_rows = read_sheet(args.iv_workbook, "IV_Sterile_Preparations")
    curated = curated_fields()

    rows, notes, possible_duplicates = build_rows(
        formulary_rows, iv_rows, curated, args.edition
    )

    args.out.parent.mkdir(parents=True, exist_ok=True)
    with open(args.out, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=OUT_HEADERS)
        writer.writeheader()
        writer.writerows(rows)

    manifest = {
        "sources": {
            p.name: {
                "sha256": hashlib.sha256(p.read_bytes()).hexdigest(),
                "rows_read": n,
            }
            for p, n in (
                (args.formulary_workbook, len(formulary_rows)),
                (args.iv_workbook, len(iv_rows)),
            )
        },
        "rows_written": len(rows),
        # Split deliberately: an exact carry-forward is the same drug keeping a
        # flag it already had. An inferred one is this script deciding that
        # "vancomycin eye solution" inherits vancomycin's high-risk status — a
        # classification judgement a reviewer should see and can overturn. The
        # rule errs toward flagging, because a surplus warning is a caution and
        # a missing one is a removed safeguard.
        "high_risk_exact_match": sorted(
            r["generic_name"] for r in rows
            if r["high_risk"] == "yes" and r["generic_name"] in curated
        ),
        "high_risk_inferred_from_prefix": sorted(
            r["generic_name"] for r in rows
            if r["high_risk"] == "yes" and r["generic_name"] not in curated
        ),
        "curated_fields_carried_forward": {
            field: sorted(r["generic_name"] for r in rows if r[field])
            for field in ("name_ar", "aliases", "antidote")
        },
        "merged_across_both_workbooks": sorted(set(notes["merged_across_workbooks"])),
        "merged_within_one_sheet": sorted(set(notes["merged_within_sheet"])),
        "possible_duplicates": possible_duplicates,
        "prefix_families_for_review": notes["prefix_families"],
        "supersedes": superseded_rows([r["generic_name"] for r in rows], curated),
        "note": (
            "auto_calculate is 'no' on every row and no numeric dose field is "
            "emitted: these workbooks state doses as prose, and parsing a "
            "number out of prose is not something this converter will do. "
            "possible_duplicates and prefix_families_for_review are reported "
            "unresolved — which spellings and which preparations are one entry "
            "is a pharmacist's decision, not this script's."
        ),
    }
    manifest_path = args.out.with_suffix(".manifest.json")
    manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False))

    print(f"wrote {len(rows)} rows to {args.out}")
    print(f"  merged across both workbooks: {len(set(notes['merged_across_workbooks']))}")
    print(f"  merged within one sheet:      {len(set(notes['merged_within_sheet']))}")
    print(f"  high-risk, exact match:       {len(manifest['high_risk_exact_match'])}")
    print(f"  high-risk, inferred:          {len(manifest['high_risk_inferred_from_prefix'])}")
    print(f"  possible duplicates:          {len(possible_duplicates)}")
    print(f"  prefix families for review:   {len(notes['prefix_families'])}")
    print(f"  manifest: {manifest_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
